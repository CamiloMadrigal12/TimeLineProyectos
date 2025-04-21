import pandas as pd
import time
import os
import re
from datetime import datetime
import shutil
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter

# Rutas
ruta_excel_proyecto = r"C:\Users\mondr\OneDrive\proyectos.xlsx"
ruta_excel_radicados = r"C:\Users\mondr\OneDrive\radicados.xlsx"
ruta_html_principal = "formulacion.html"
ruta_diagnosticos = "diagnosticos"
ruta_html_radicado = "radicados.html"
ruta_diagnostico = "diagnostico_radicado"

# Definir ruta_backup
ruta_backup = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backup_radicados.xlsx")
# Archivo para guardar la fecha de la última actualización
ruta_ultima_actualizacion = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ultima_actualizacion.txt")

def obtener_dias_transcurridos():
    """
    Calcula cuántos días han pasado desde la última actualización.
    Si es la primera vez que se ejecuta, retorna 0.
    """
    hoy = datetime.now().date()
    
    if os.path.exists(ruta_ultima_actualizacion):
        try:
            with open(ruta_ultima_actualizacion, 'r') as f:
                fecha_str = f.read().strip()
                ultima_fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                dias_transcurridos = (hoy - ultima_fecha).days
                return max(0, dias_transcurridos)  # Asegurarse de que no sea negativo
        except Exception as e:
            print(f"⚠️ Error al leer la fecha de última actualización: {e}")
            return 0
    else:
        return 0  # Primera ejecución

def guardar_fecha_actualizacion():
    """
    Guarda la fecha actual como la última fecha de actualización.
    """
    try:
        hoy = datetime.now().date().strftime('%Y-%m-%d')
        with open(ruta_ultima_actualizacion, 'w') as f:
            f.write(hoy)
        print(f"✅ Fecha de actualización guardada: {hoy}")
    except Exception as e:
        print(f"⚠️ Error al guardar la fecha de actualización: {e}")

def extraer_dias(valor):
    """
    Extrae el número de días de un valor que puede ser un número o una cadena.
    """
    try:
        # Si es una cadena que contiene "días", extraer solo el número
        if isinstance(valor, str):
            if "días" in valor:
                return int(valor.replace("días", "").strip())
            else:
                return int(valor.strip())
        # Si es un número, devolverlo directamente
        elif isinstance(valor, (int, float)):
            return int(valor)
        else:
            return 0
    except (ValueError, TypeError):
        return 0  # Si hay un error, devolver 0

def calcular_semaforo_html(dias):
    """
    Convierte los días en un semáforo de colores según los requisitos:
    - 0-50 días: verde
    - 51-80 días: amarillo
    - 81+ días: rojo
    
    Retorna HTML con un span que contiene el número de días y un círculo de color.
    """
    try:
        # Convertir a entero si es posible
        dias = int(dias)
    except (ValueError, TypeError):
        dias = 0  # Si hay un error, asignamos 0 días por defecto

    # Definir la clase CSS según los días
    if dias >= 81:
        clase = "semaforo-rojo"  # Rojo
    elif dias >= 51:
        clase = "semaforo-amarillo"  # Amarillo
    else:
        clase = "semaforo-verde"  # Verde

    # Retornar HTML con el span y la clase correspondiente
    return f'{dias} <span class="{clase}"></span>'

def verificar_cambios_observacion(df_actual, respetar_dias_existentes=False):
    """
    Verifica si ha habido cambios en la columna de observación y
    reinicia el contador de días SOLO para los proyectos específicos
    cuya observación ha cambiado.
    
    Args:
        df_actual: DataFrame con los datos actuales
        respetar_dias_existentes: Si es True, respeta los días existentes en el Excel
                                 y no los actualiza en la primera ejecución
    """
    # Calcular días transcurridos desde la última actualización
    dias_transcurridos = obtener_dias_transcurridos()
    print(f"📅 Han pasado {dias_transcurridos} días desde la última actualización.")
    
    # Identificar la columna de observación (puede tener diferentes nombres)
    columna_observacion = None
    posibles_nombres = ['OBSERVACION', 'Observacion', 'observacion', 'OBSERVACIÓN', 'Observación', 'observación', 'OBSERVACIONES', ]
    for nombre in posibles_nombres:
        if nombre in df_actual.columns:
            columna_observacion = nombre
            break
    
    if columna_observacion is None:
        print("⚠️ No se encontró la columna de observación. No se podrán detectar cambios.")
        return df_actual
    
    # Verificar si existe un respaldo anterior
    if os.path.exists(ruta_backup):
        try:
            # Cargar el DataFrame del respaldo
            df_anterior = pd.read_excel(ruta_backup, sheet_name="radicados", engine="openpyxl")
            
            # Iterar por cada fila del DataFrame actual
            for index, row in df_actual.iterrows():
                ci = row["#CI"]
                observacion_actual = str(row.get(columna_observacion, ""))
                
                # Buscar la misma fila en el DataFrame anterior
                fila_anterior = df_anterior[df_anterior["#CI"] == ci]
                if not fila_anterior.empty:
                    # Verificar si la columna de observación existe en el DataFrame anterior
                    if columna_observacion in df_anterior.columns:
                        observacion_anterior = str(fila_anterior.iloc[0].get(columna_observacion, ""))
                        
                        # Si la observación ha cambiado y no está vacía, reiniciar el contador de días
                        # SOLO para este proyecto específico
                        if observacion_actual != observacion_anterior and observacion_actual.strip():
                            print(f"✅ Se detectó una nueva observación para el CI {ci}. Reiniciando contador de días.")
                            df_actual.at[index, "DIAS DESDE LA ULTIMA OBSERVACION"] = 0
                        else:
                            # Si no ha cambiado, actualizar el contador sumando los días transcurridos
                            dias_actuales = extraer_dias(row.get("DIAS DESDE LA ULTIMA OBSERVACION", 0))
                            df_actual.at[index, "DIAS DESDE LA ULTIMA OBSERVACION"] = dias_actuales + dias_transcurridos
                    else:
                        # Si no existe la columna en el respaldo, solo actualizar el contador
                        dias_actuales = extraer_dias(row.get("DIAS DESDE LA ULTIMA OBSERVACION", 0))
                        df_actual.at[index, "DIAS DESDE LA ULTIMA OBSERVACION"] = dias_actuales + dias_transcurridos
                else:
                    # Si es un nuevo proyecto, iniciar el contador en 0
                    print(f"✅ Nuevo proyecto detectado: CI {ci}. Iniciando contador de días.")
                    df_actual.at[index, "DIAS DESDE LA ULTIMA OBSERVACION"] = 0
        except Exception as e:
            print(f"⚠️ Error al verificar cambios en observaciones: {e}")
            # En caso de error, actualizar todos los contadores
            for index, row in df_actual.iterrows():
                dias_actuales = extraer_dias(row.get("DIAS DESDE LA ULTIMA OBSERVACION", 0))
                df_actual.at[index, "DIAS DESDE LA ULTIMA OBSERVACION"] = dias_actuales + dias_transcurridos
    else:
        # Si no hay respaldo, decidir qué hacer según el parámetro respetar_dias_existentes
        if respetar_dias_existentes:
            print("📊 Primera ejecución: Respetando los días existentes en el Excel original.")
            # No modificamos los valores existentes, solo creamos el backup
        else:
            # Si no hay respaldo, actualizar todos los contadores
            print("📊 Primera ejecución: Actualizando contadores con días transcurridos.")
            for index, row in df_actual.iterrows():
                dias_actuales = extraer_dias(row.get("DIAS DESDE LA ULTIMA OBSERVACION", 0))
                df_actual.at[index, "DIAS DESDE LA ULTIMA OBSERVACION"] = dias_actuales + dias_transcurridos
    
    # Guardar una copia del DataFrame actual como respaldo para la próxima comparación
    try:
        # Crear un nuevo archivo Excel solo para el backup
        with pd.ExcelWriter(ruta_backup, engine="openpyxl") as writer:
            df_actual.to_excel(writer, sheet_name="radicados", index=False)
        print(f"✅ Backup guardado en: {ruta_backup}")
        
        # Guardar la fecha actual como última actualización
        guardar_fecha_actualizacion()
    except Exception as e:
        print(f"⚠️ Error al guardar backup: {e}")
    
    return df_actual

def aplicar_estilos_excel(workbook, worksheet, df):
    """
    Aplica estilos profesionales a una hoja de Excel.
    
    Args:
        workbook: Objeto de libro de trabajo de openpyxl
        worksheet: Objeto de hoja de trabajo de openpyxl
        df: DataFrame con los datos
    """
    # Definir estilos
    # Borde estándar
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Estilo para encabezados
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Estilo para filas alternas
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Estilo para semáforo
    verde_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    amarillo_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    rojo_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Aplicar estilos a los encabezados
    for col_idx, column in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = header_alignment
    
    # Aplicar estilos a las celdas de datos
    for row_idx, row in enumerate(df.values, 2):
        # Aplicar fondo alterno a filas
        row_fill = alt_fill if row_idx % 2 == 0 else None
        
        for col_idx, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            
            # Aplicar fondo alterno
            if row_fill:
                cell.fill = row_fill
            
            # Centrar el texto
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar formato condicional a la columna de días
            if df.columns[col_idx-1] == "DIAS DESDE LA ULTIMA OBSERVACION" and isinstance(value, (int, float)):
                if value >= 81:
                    cell.fill = rojo_fill
                    cell.font = Font(color='9C0006')
                elif value >= 51:
                    cell.fill = amarillo_fill
                    cell.font = Font(color='9C5700')
                else:
                    cell.fill = verde_fill
                    cell.font = Font(color='006100')
    
    # Ajustar el ancho de las columnas
    for col_idx, column in enumerate(df.columns, 1):
        column_width = max(
            len(str(column)),
            max(len(str(value)) for value in df[column].fillna('')) if len(df) > 0 else 0
        )
        # Añadir un poco de espacio extra
        column_width = min(column_width + 4, 50)  # Limitar a 50 caracteres máximo
        worksheet.column_dimensions[get_column_letter(col_idx)].width = column_width
    
    # Congelar la primera fila (encabezados)
    worksheet.freeze_panes = 'A2'

def actualizar_hoja_excel_seguro(df_actualizado, ruta_excel, nombre_hoja):
    """
    Actualiza una hoja específica en un archivo Excel sin afectar las demás hojas.
    Método seguro que lee todas las hojas, actualiza una y guarda todas.
    Aplica estilos profesionales a la hoja actualizada.
    
    Args:
        df_actualizado: DataFrame con los datos actualizados
        ruta_excel: Ruta al archivo Excel
        nombre_hoja: Nombre de la hoja a actualizar
    """
    try:
        # Verificar si el archivo existe
        if not os.path.exists(ruta_excel):
            print(f"⚠️ El archivo {ruta_excel} no existe. Creando nuevo archivo.")
            with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
                df_actualizado.to_excel(writer, sheet_name=nombre_hoja, index=False)
                workbook = writer.book
                worksheet = writer.sheets[nombre_hoja]
                aplicar_estilos_excel(workbook, worksheet, df_actualizado)
            return True
        
        # Crear un archivo temporal para la copia de seguridad
        temp_backup = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
        shutil.copy2(ruta_excel, temp_backup)
        print(f"✅ Copia de seguridad temporal creada en: {temp_backup}")
        
        try:
            # Leer todas las hojas del archivo Excel
            excel_file = pd.ExcelFile(ruta_excel, engine="openpyxl")
            sheet_names = excel_file.sheet_names
            
            # Crear un diccionario para almacenar todos los DataFrames
            all_sheets = {}
            
            # Leer todas las hojas
            for sheet in sheet_names:
                all_sheets[sheet] = pd.read_excel(ruta_excel, sheet_name=sheet, engine="openpyxl")
            
            # Actualizar la hoja específica
            all_sheets[nombre_hoja] = df_actualizado
            
            # Guardar todas las hojas en un nuevo archivo Excel
            with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
                for sheet_name, df in all_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Aplicar estilos solo a la hoja que estamos actualizando
                    if sheet_name == nombre_hoja:
                        workbook = writer.book
                        worksheet = writer.sheets[sheet_name]
                        aplicar_estilos_excel(workbook, worksheet, df)
            
            print(f"✅ Hoja '{nombre_hoja}' actualizada con éxito en {ruta_excel}")
            
            # Eliminar la copia de seguridad temporal
            os.remove(temp_backup)
            return True
            
        except Exception as e:
            print(f"❌ Error al actualizar la hoja '{nombre_hoja}': {e}")
            # Restaurar desde la copia de seguridad
            shutil.copy2(temp_backup, ruta_excel)
            print(f"✅ Archivo restaurado desde la copia de seguridad")
            # Eliminar la copia de seguridad temporal
            os.remove(temp_backup)
            return False
    
    except Exception as e:
        print(f"❌ Error general al actualizar el Excel: {e}")
        return False

def generar_html_principal():
    if not os.path.exists(ruta_excel_proyecto):
        print("❌ ERROR: El archivo de proyectos no existe o no está accesible.")
        return

    intentos = 5
    while intentos > 0:
        try:
            time.sleep(1)
            df_proyectos = pd.read_excel(ruta_excel_proyecto, sheet_name="proyectos", engine="openpyxl")
            break  
        except PermissionError:
            print(f"⚠️ Archivo en uso, esperando {intentos} segundos para intentar de nuevo...")
            time.sleep(1)
            intentos -= 1
        except Exception as e:
            print(f"❌ ERROR al leer el archivo Excel de proyectos: {e}")
            return

    if "#CI" not in df_proyectos.columns:
        print("❌ ERROR: La columna '#CI' no existe en el archivo Excel de proyectos.")
        return

    # Crear una copia del DataFrame para modificar
    df_proyectos_html = df_proyectos.copy()
    
    # Ocultar las columnas especificadas
    columnas_a_ocultar = ['ID MGA', 'BPIN', 'RADICADO', 'FECHA', 'VALOR', 'RADICADO VIABILIDAD', 'FECHA VIABILIDAD', 'OBSERVACIONES']
    for col in columnas_a_ocultar:
        if col in df_proyectos_html.columns:
            df_proyectos_html = df_proyectos_html.drop(columns=[col])
    
    for index, row in df_proyectos_html.iterrows():
        ci_numero = row["#CI"]
        archivo_diagnostico = f"{ruta_diagnosticos}/{ci_numero}.html"
        df_proyectos_html.at[index, "#CI"] = f'<a href="{archivo_diagnostico}" target="_blank">{ci_numero}</a>'

    tabla_html = df_proyectos_html.to_html(index=False, escape=False, classes="table table-bordered table-striped")

    if os.path.exists(ruta_html_principal):
        with open(ruta_html_principal, "r", encoding="utf-8") as file:
            contenido_html = file.read()
        nuevo_html = re.sub(r"(<table.*?>).*?(</table>)", f"\\1{tabla_html}\\2", contenido_html, flags=re.DOTALL)
    else:
        nuevo_html = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>DAP</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
             <style>
                /* Estilos para el semáforo */
                .semaforo-verde, .semaforo-amarillo, .semaforo-rojo {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    border-radius: 50%;
                    margin-left: 5px;
                    vertical-align: middle;
                }}
                .semaforo-verde {{
                    background-color: #28a745; /* Verde */
                }}
                .semaforo-amarillo {{
                    background-color: #ffc107; /* Amarillo */
                }}
                .semaforo-rojo {{
                    background-color: #dc3545; /* Rojo */
                }}
                /* Centrar los encabezados de la tabla */
                .table th {{
                    text-align: center;
                }}
            </style>
        </head>
        <body>
            <div id="header"></div>
            <div class="container mt-4">
                <h2 class="text-center">Proyectos en Formulación</h2>
                {tabla_html}
            </div>
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
            <script>
                document.addEventListener("DOMContentLoaded", function () {{
                    let headerPath = "https://camilomadrigal12.github.io/TimeLineProyectos/header.html";
                    fetch(headerPath)
                        .then(response => response.text())
                        .then(data => {{
                            document.getElementById("header").innerHTML = data;
                        }})
                        .catch(error => console.error("Error cargando el header:", error));
                }});
            </script>
        </body>
        </html>
        """

    with open(ruta_html_principal, "w", encoding="utf-8") as file:
        file.write(nuevo_html)

    print("✅ Archivo HTML principal actualizado con éxito.")

def generar_html_radicado(respetar_dias_existentes=False):
    if not os.path.exists(ruta_excel_radicados):
        print("❌ ERROR: El archivo de radicados no existe o no está accesible.")
        return

    intentos = 5
    while intentos > 0:
        try:
            time.sleep(1)
            df_radicados = pd.read_excel(ruta_excel_radicados, sheet_name="radicados", engine="openpyxl")
            break  
        except PermissionError:
            print(f"⚠️ Archivo en uso, esperando {intentos} segundos para intentar de nuevo...")
            time.sleep(1)
            intentos -= 1
        except Exception as e:
            print(f"❌ ERROR al leer el archivo Excel de radicados: {e}")
            return

    if "#CI" not in df_radicados.columns:
        print("❌ ERROR: La columna '#CI' no existe en el archivo Excel de radicados.")
        return

    # Verificar cambios en las observaciones y reiniciar el contador si es necesario
    df_radicados = verificar_cambios_observacion(df_radicados, respetar_dias_existentes)
    
    # SOLUCIÓN CORREGIDA: Actualizar solo la hoja "radicados" sin afectar las demás hojas
    actualizar_hoja_excel_seguro(df_radicados, ruta_excel_radicados, "radicados")

    # Crear una copia del DataFrame para modificar para HTML
    df_radicados_html = df_radicados.copy()
    
    # Convertir la columna de días a numérica
    if 'DIAS DESDE LA ULTIMA OBSERVACION' in df_radicados_html.columns:
        df_radicados_html['DIAS DESDE LA ULTIMA OBSERVACION'] = pd.to_numeric(
            df_radicados_html['DIAS DESDE LA ULTIMA OBSERVACION'], errors='coerce').fillna(0).astype(int)
        
        # Aplicar la función de semáforo
        df_radicados_html['DIAS DESDE LA ULTIMA OBSERVACION'] = df_radicados_html['DIAS DESDE LA ULTIMA OBSERVACION'].apply(calcular_semaforo_html)

    for index, row in df_radicados_html.iterrows():
        ci_numero = row["#CI"]
        archivo_diagnostico_radicado = f"{ruta_diagnostico}/{ci_numero}.html"
        df_radicados_html.at[index, "#CI"] = f'<a href="{archivo_diagnostico_radicado}" target="_blank">{ci_numero}</a>'

    # IMPORTANTE: Crear una copia del DataFrame sin la columna de observación para mostrar en la web
    # Verificar todos los posibles nombres de la columna (mayúsculas, minúsculas, etc.)
    columnas_a_ocultar = ['OBSERVACION', 'Observacion', 'observacion', 'OBSERVACIÓN', 'Observación', 'observación', 'OBSERVACIONES']
    df_radicados_web = df_radicados_html.copy()
    for col in columnas_a_ocultar:
        if col in df_radicados_web.columns:
            df_radicados_web = df_radicados_web.drop(columns=[col])
    
    # Asegurarse de que la columna BPIN esté visible (no se oculta)
    # No es necesario hacer nada especial, ya que solo ocultamos las columnas específicas
    
    # Imprimir las columnas para verificar que la columna de observación se haya eliminado
    print("Columnas en el DataFrame web:", df_radicados_web.columns.tolist())
    
    # Usar df_radicados_web para generar la tabla HTML
    tabla_html = df_radicados_web.to_html(index=False, escape=False, classes="table table-bordered table-striped")
    tabla_html = tabla_html.replace("<th", "<th style='text-align: center;'")

    if os.path.exists(ruta_html_radicado):
        with open(ruta_html_radicado, "r", encoding="utf-8") as file:
            contenido_html = file.read()
        # Verificar si ya existen los estilos del semáforo
        if ".semaforo-verde" not in contenido_html:
            # Agregar estilos si no existen
            contenido_html = contenido_html.replace("</head>", """
            <style>
                /* Estilos para el semáforo */
                .semaforo-verde, .semaforo-amarillo, .semaforo-rojo {
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    border-radius: 50%;
                    margin-left: 5px;
                    vertical-align: middle;
                }
                .semaforo-verde {
                    background-color: #28a745; /* Verde */
                }
                .semaforo-amarillo {
                    background-color: #ffc107; /* Amarillo */
                }
                .semaforo-rojo {
                    background-color: #dc3545; /* Rojo */
                }
                /* Centrar los encabezados de la tabla */
                .table th {
                    text-align: center;
                }
            </style>
            </head>""")
        nuevo_html = re.sub(r"(<table.*?>).*?(</table>)", f"\\1{tabla_html}\\2", contenido_html, flags=re.DOTALL)
    else:
        nuevo_html = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>DAP</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
            <style>
                /* Estilos para el semáforo */
                .semaforo-verde, .semaforo-amarillo, .semaforo-rojo {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    border-radius: 50%;
                    margin-left: 5px;
                    vertical-align: middle;
                }}
                .semaforo-verde {{
                    background-color: #28a745; /* Verde */
                }}
                .semaforo-amarillo {{
                    background-color: #ffc107; /* Amarillo */
                }}
                .semaforo-rojo {{
                    background-color: #dc3545; /* Rojo */
                }}
                /* Centrar los encabezados de la tabla */
                .table th {{
                    text-align: center;
                }}
            </style>
        </head>
        <body>
            <div id="header"></div>
            <div class="container mt-4">
                <h2 class="text-center">Proyecto Radicados</h2>
                {tabla_html}
            </div>
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
            <script>
                document.addEventListener("DOMContentLoaded", function () {{
                    let headerPath = "https://camilomadrigal12.github.io/TimeLineProyectos/header.html";
                    fetch(headerPath)
                        .then(response => response.text())
                        .then(data => {{
                            document.getElementById("header").innerHTML = data;
                        }})
                        .catch(error => console.error("Error cargando el header:", error));
                }});
            </script>
        </body>
        </html>
        """

    with open(ruta_html_radicado, "w", encoding="utf-8") as file:
        file.write(nuevo_html)

    print("✅ Archivo HTML de radicados actualizado con éxito.")

def generar_timeline_html(proceso):
    if proceso is None:
        return ""
        
    fases = [
        'En Diagnóstico', 'Planteamiento Problema', 'Alternativas Solución', 
        'Listado de Requisitos', 'Estudio y Documentación', 'Proceso en Costos', 
        'Certificados y Firmas', 'Proceso en Radicación'
    ]
    
    fase_actual = proceso.get('Proceso Actual', '')
    estado_actual = proceso.get('Estado', '')
    fase_numero = int(proceso.get('FASE DEL PROCESO', 0))
    
    timeline_html = ""
    for i, fase in enumerate(fases, start=1):
        completed = ""
        checked = ""
        checkmark = ""
        
        if i < fase_numero:
            completed = "completed"
            checked = "checked"
            checkmark = '<span class="checkmark">✔</span>'
        elif i == fase_numero:
            if estado_actual == "Completado":
                completed = "completed"
                checked = "checked"
                checkmark = '<span class="checkmark">✔</span>'
            elif estado_actual == "En Proceso":
                completed = "completed"
        
        timeline_html += f"""
        <div class="timeline-step {completed} {checked}">
            <div class="dot"></div>
            <h3>{fase} {checkmark}</h3>
        </div>
        """
    
    return timeline_html

def generar_html_diagnosticos():
    if not os.path.exists(ruta_excel_proyecto):
        print("❌ ERROR: Uno o ambos archivos Excel no existen o no están accesibles.")
        return

    try:
        df_proyectos = pd.read_excel(ruta_excel_proyecto, sheet_name="proyectos", engine="openpyxl")
        df_procesos = pd.read_excel(ruta_excel_proyecto, sheet_name="procesos", engine="openpyxl")
        
        # Aplicar estilos a la hoja de procesos
        with pd.ExcelWriter(ruta_excel_proyecto, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            workbook = writer.book
            if "procesos" in workbook.sheetnames:
                worksheet = workbook["procesos"]
                aplicar_estilos_excel(workbook, worksheet, df_procesos)
                
    except Exception as e:
        print(f"❌ ERROR al leer los archivos Excel: {e}")
        return

    if "#CI" not in df_proyectos.columns or "#CI" not in df_procesos.columns:
        print("❌ ERROR: La columna '#CI' no existe en uno o ambos archivos Excel.")
        return

    if not os.path.exists(ruta_diagnosticos):
        os.makedirs(ruta_diagnosticos)

    for _, proyecto in df_proyectos.iterrows():
        ci = proyecto["#CI"]
        proceso = df_procesos[df_procesos["#CI"] == ci].iloc[0] if not df_procesos[df_procesos["#CI"] == ci].empty else None

        timeline_html = generar_timeline_html(proceso) if proceso is not None else ""

        # Solución para el problema de rutas relativas en subcarpetas
        html_content = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Diagnóstico Proyecto #{ci}</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
            <!-- Agregar base href para corregir rutas relativas -->
            <base href="../">
            <style>
            .semaforo-verde, .semaforo-amarillo, .semaforo-rojo {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    border-radius: 50%;
                    margin-left: 5px;
                    vertical-align: middle;
                }}
                .semaforo-verde {{
                    background-color: #28a745; /* Verde */
                }}
                .semaforo-amarillo {{
                    background-color: #ffc107; /* Amarillo */
                }}
                .semaforo-rojo {{
                    background-color: #dc3545; /* Rojo */
                }}
                body {{
                    font-family: Arial, sans-serif;
                    background-color: #f4f4f9;
                    margin: 0;
                    padding: 0;
                }}
                .content {{
                    padding: 80px 20px;
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    justify-content: center;
                    min-height: calc(100vh - 80px);
                    text-align: center;
                }}
                .title {{
                    font-size: 24px;
                    font-weight: bold;
                    color: #333;
                    margin-bottom: 20px;
                }}
                .timeline {{
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    width: 90%;
                    max-width: 1000px;
                    position: relative;
                    margin-top: 40px;
                }}
                .timeline::before {{
                    content: '';
                    position: absolute;
                    top: 31%;
                    left: 0;
                    right: 0;
                    height: 4px;
                    background: #0078d4;
                    transform: translateY(-50%);
                    z-index: 1;
                }}
                .timeline-step {{
                    position: relative;
                    flex: 1;
                    text-align: center;
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    padding: 0 10px;
                }}
                .timeline-step .dot {{
                    width: 15px;
                    height: 15px;
                    background: #0078d4;
                    border: 3px solid #fff;
                    border-radius: 50%;
                    z-index: 2;
                }}
                .timeline-step.completed .dot {{
                    background: #0078d4;
                }}
                .timeline-step.checked .dot {{
                    background: #28a745;
                }}
                .checkmark {{
                    display: inline-block;
                    margin-top: 10px;
                    color: #28a745;
                    font-size: 18px;
                }}
                .timeline-step h3 {{
                    font-size: 16px;
                    margin-top: 15px;
                }}
                @media (max-width: 768px) {{
                    .timeline {{
                        flex-direction: column;
                        align-items: center;
                    }}
                    .timeline::before {{
                        width: 4px;
                        height: 100%;
                        left: 50%;
                        transform: translateX(-50%);
                    }}
                    .timeline-step {{
                        flex-direction: row;
                        text-align: left;
                        justify-content: start;
                        width: 100%;
                        padding: 10px 0;
                    }}
                    .timeline-step .dot {{
                        width: 12px;
                        height: 12px;
                        margin-bottom: 0;
                        margin-right: 10px;
                    }}
                    .timeline-step h3 {{
                        font-size: 14px;
                    }}
                }}
            </style>
        </head>
        <body>
            <div id="header"></div>
            <div class="content">
                <div class="title">Diagrama de Procesos - Proyecto #{ci}</div>
                <div class="card mb-4">
                    <div class="card-body">
                        <h5 class="card-title">{proyecto['NOMBRE DEL PROYECTO']}</h5>
                        <p class="card-text"><strong>Dependencia:</strong> {proyecto.get('DEPENDENCIA', '')}</p>
                        <p class="card-text"><strong>Fase del Proyecto:</strong> {proyecto['FASE DEL PROYECTO']}</p>
                        <p class="card-text"><strong>Entidad Destino para Presentacion:</strong> {proyecto['ENTIDAD DESTINO PARA PRESENTACION']}</p>
                    </div>
                </div>
                <div class="timeline">
                    {timeline_html}
                </div>
            </div>
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
            <script>
                document.addEventListener("DOMContentLoaded", function () {{
                    let headerPath = "https://camilomadrigal12.github.io/TimeLineProyectos/header.html";
                    fetch(headerPath)
                        .then(response => {{
                            if (!response.ok) {{
                                throw new Error("No se pudo cargar el header.");
                            }}
                            return response.text();
                        }})
                        .then(data => {{
                            document.getElementById("header").innerHTML = data;

                            setTimeout(() => {{
                                let logo = document.querySelector("#header #logo");
                                if (logo) {{
                                    logo.src = logo.getAttribute("data-src");
                                }}

                                let links = document.querySelectorAll("#header a[data-href]");
                                links.forEach(link => {{
                                    link.setAttribute("href", link.getAttribute("data-href"));
                                }});
                            }}, 500);
                        }})
                        .catch(error => console.error("Error cargando el header:", error));
                }});
            </script>
        </body>
        </html>
        """

        with open(f"{ruta_diagnosticos}/{ci}.html", "w", encoding="utf-8") as file:
            file.write(html_content)

    print("✅ Archivos HTML de diagnóstico actualizados con éxito.")

def generar_html_diagnostico_radicado():
    if not os.path.exists(ruta_excel_radicados):
        print("❌ El archivo Excel de radicados no existe o no está accesible.")
        return

    try:
        df_radicados = pd.read_excel(ruta_excel_radicados, sheet_name="radicados", engine="openpyxl")
        df_diagnostico = pd.read_excel(ruta_excel_radicados, sheet_name="diagnostico", engine="openpyxl")
        
        # Aplicar estilos a la hoja de diagnóstico
        with pd.ExcelWriter(ruta_excel_radicados, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            workbook = writer.book
            if "diagnostico" in workbook.sheetnames:
                worksheet = workbook["diagnostico"]
                aplicar_estilos_excel(workbook, worksheet, df_diagnostico)
        
        # Ocultar columnas G y H desde la fila 8 a 15 en la hoja de diagnóstico
        # Esto no afecta la visualización en HTML, solo es para referencia
        print("ℹ️ Las columnas G y H desde la fila 8 a 15 en la hoja de diagnóstico serán ocultadas en la visualización.")
        
    except Exception as e:
        print(f"❌ ERROR al leer los archivos Excel: {e}")
        return

    if "#CI" not in df_radicados.columns or "#CI" not in df_diagnostico.columns:
        print("❌ ERROR: La columna '#CI' no existe en uno o ambos archivos Excel.")
        return

    if not os.path.exists(ruta_diagnostico):
        os.makedirs(ruta_diagnostico)

    # Verificar cambios en las observaciones y reiniciar el contador si es necesario
    df_radicados = verificar_cambios_observacion(df_radicados)

    # Convertir la columna de días a numérica
    if 'DIAS DESDE LA ULTIMA OBSERVACION' in df_radicados.columns:
        df_radicados['DIAS DESDE LA ULTIMA OBSERVACION'] = pd.to_numeric(
            df_radicados['DIAS DESDE LA ULTIMA OBSERVACION'], errors='coerce').fillna(0).astype(int)
        
        # Aplicar la función de semáforo
        df_radicados_con_semaforo = df_radicados.copy()
        df_radicados_con_semaforo['DIAS_CON_SEMAFORO'] = df_radicados['DIAS DESDE LA ULTIMA OBSERVACION'].apply(calcular_semaforo_html)

    for _, radicado in df_radicados.iterrows():
        ci = radicado["#CI"]
        diagnostico = df_diagnostico[df_diagnostico["#CI"] == ci].iloc[0] if not df_diagnostico[df_diagnostico["#CI"] == ci].empty else None

        timeline_html = generar_timeline_html(diagnostico) if diagnostico is not None else ""
        
        # Obtener el valor con semáforo para este radicado
        dias_con_semaforo = ""
        if 'DIAS_CON_SEMAFORO' in df_radicados_con_semaforo.columns:
            fila = df_radicados_con_semaforo[df_radicados_con_semaforo["#CI"] == ci]
            if not fila.empty:
                dias_con_semaforo = fila.iloc[0]['DIAS_CON_SEMAFORO']

        # Solución para el problema de rutas relativas en subcarpetas
        html_content = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Diagnóstico Radicado #{ci}</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
            <!-- Agregar base href para corregir rutas relativas -->
            <base href="../">
            <style>
            .semaforo-verde, .semaforo-amarillo, .semaforo-rojo {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    border-radius: 50%;
                    margin-left: 5px;
                    vertical-align: middle;
                }}
                .semaforo-verde {{
                    background-color: #28a745; /* Verde */
                }}
                .semaforo-amarillo {{
                    background-color: #ffc107; /* Amarillo */
                }}
                .semaforo-rojo {{
                    background-color: #dc3545; /* Rojo */
                }}
                body {{
                    font-family: Arial, sans-serif;
                    background-color: #f4f4f9;
                    margin: 0;
                    padding: 0;
                }}
                .content {{
                    padding: 80px 20px;
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    justify-content: center;
                    min-height: calc(100vh - 80px);
                    text-align: center;
                }}
                .title {{
                    font-size: 24px;
                    font-weight: bold;
                    color: #333;
                    margin-bottom: 20px;
                }}
                .timeline {{
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    width: 90%;
                    max-width: 1000px;
                    position: relative;
                    margin-top: 40px;
                }}
                .timeline::before {{
                    content: '';
                    position: absolute;
                    top: 31%;
                    left: 0;
                    right: 0;
                    height: 4px;
                    background: #0078d4;
                    transform: translateY(-50%);
                    z-index: 1;
                }}
                .timeline-step {{
                    position: relative;
                    flex: 1;
                    text-align: center;
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    padding: 0 10px;
                }}
                .timeline-step .dot {{
                    width: 15px;
                    height: 15px;
                    background: #0078d4;
                    border: 3px solid #fff;
                    border-radius: 50%;
                    z-index: 2;
                }}
                .timeline-step.completed .dot {{
                    background: #0078d4;
                }}
                .timeline-step.checked .dot {{
                    background: #28a745;
                }}
                .checkmark {{
                    display: inline-block;
                    margin-top: 10px;
                    color: #28a745;
                    font-size: 18px;
                }}
                .timeline-step h3 {{
                    font-size: 16px;
                    margin-top: 15px;
                }}
                @media (max-width: 768px) {{
                    .timeline {{
                        flex-direction: column;
                        align-items: center;
                    }}
                    .timeline::before {{
                        width: 4px;
                        height: 100%;
                        left: 50%;
                        transform: translateX(-50%);
                    }}
                    .timeline-step {{
                        flex-direction: row;
                        text-align: left;
                        justify-content: start;
                        width: 100%;
                        padding: 10px 0;
                    }}
                    .timeline-step .dot {{
                        width: 12px;
                        height: 12px;
                        margin-bottom: 0;
                        margin-right: 10px;
                    }}
                    .timeline-step h3 {{
                        font-size: 14px;
                    }}
                }}
            </style>
        </head>
        <body>
            <div id="header"></div>
            <div class="content">
                <div class="title">Diagrama de Radicados - Proyecto #{ci}</div>
                <div class="card mb-4">
                    <div class="card-body">
                        <p class="card-text"><strong>ID MGA:</strong> {radicado['ID MGA']}</p>
                        <p class="card-text"><strong>BPIN:</strong> {radicado.get('BPIN', '')}</p>
                        <h5 class="card-title">{radicado['NOMBRE DEL PROYECTO']}</h5>
                        <p class="card-text"><strong>Fase del Proyecto:</strong> {radicado['FASE DEL PROYECTO']}</p>
                        <p class="card-text"><strong>Entidad Destino para presentacion:</strong> {radicado['ENTIDAD DESTINO PARA PRESENTACION']}</p>
                        <p class="card-text"><strong># Radicado:</strong> {radicado['# RADICADO']}</p>
                        <p class="card-text"><strong>Fecha Radicado:</strong> {radicado['FECHA RADICADO']}</p>
                        <p class="card-text"><strong>Valor:</strong> {radicado['VALOR']}</p>
                        <p class="card-text"><strong>Días desde la Última Observación:</strong> {dias_con_semaforo}</p>
                    </div>
                </div>
                <div class="timeline">
                    {timeline_html}
                </div>
            </div>
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
            <script>
                document.addEventListener("DOMContentLoaded", function () {{
                    let headerPath = "https://camilomadrigal12.github.io/TimeLineProyectos/header.html";
                    fetch(headerPath)
                        .then(response => {{
                            if (!response.ok) {{
                                throw new Error("No se pudo cargar el header.");
                            }}
                            return response.text();
                        }})
                        .then(data => {{
                            document.getElementById("header").innerHTML = data;

                            setTimeout(() => {{
                                let logo = document.querySelector("#header #logo");
                                if (logo) {{
                                    logo.src = logo.getAttribute("data-src");
                                }}

                                let links = document.querySelectorAll("#header a[data-href]");
                                links.forEach(link => {{
                                    link.setAttribute("href", link.getAttribute("data-href"));
                                }});
                            }}, 500);
                        }})
                        .catch(error => console.error("Error cargando el header:", error));
                }});
            </script>
        </body>
        </html>
        """

        with open(f"{ruta_diagnostico}/{ci}.html", "w", encoding="utf-8") as file:
            file.write(html_content)

    print("✅ Archivos HTML de diagnóstico de radicados actualizados con éxito.")

def generar_html():
    generar_html_principal()
    generar_html_radicado()
    generar_html_diagnosticos()
    generar_html_diagnostico_radicado()

if __name__ == "__main__":
    try:
        generar_html()
    except Exception as e:
        import traceback
        print(f"❌ ERROR GENERAL: {e}")
        print(traceback.format_exc())
